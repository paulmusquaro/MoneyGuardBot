import asyncio
import datetime
import io
import logging
import os
import uuid

import aiohttp
import openpyxl
from aiogram import Bot, Dispatcher, types
from aiogram.filters import CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, Message, ReplyKeyboardRemove
from aiogram.utils.keyboard import ReplyKeyboardBuilder

from dotenv import load_dotenv
import os

load_dotenv()


BOT_TOKEN = os.getenv("BOT_API_TOKEN")
API_URL = 'http://localhost:8000'


logging.basicConfig(level=logging.INFO)

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()


class ExpenseState(StatesGroup):
    name = State()
    date = State()
    amount = State()


class ReportState(StatesGroup):
    start_date = State()
    end_date = State()


class DeleteExpenseState(StatesGroup):
    expense_id = State()


class EditExpenseState(StatesGroup):
    expense_id = State()
    new_name = State()
    new_amount = State()


async def get_expenses(session: aiohttp.ClientSession):
    async with session.get(f"{API_URL}/expenses") as response:
        return await response.json()


async def add_expense(session: aiohttp.ClientSession, expense):
    async with session.post(f"{API_URL}/expenses", json=expense) as response:
        return await response.json()


async def get_report(session: aiohttp.ClientSession, start_date, end_date):
    async with session.get(
        f"{API_URL}/report", params={"start_date": start_date, "end_date": end_date}
    ) as response:
        return await response.json()


async def delete_expense(session: aiohttp.ClientSession, expense_id):
    async with session.delete(f"{API_URL}/expenses/{expense_id}") as response:
        return await response.json()


async def update_expense(session: aiohttp.ClientSession, expense_id, expense):
    async with session.put(f"{API_URL}/expenses/{expense_id}", json=expense) as response:
        return await response.json()


def create_menu_keyboard():
    builder = ReplyKeyboardBuilder()
    builder.button(text="Додати статтю витрат")
    builder.button(text="Отримати звіт витрат за період")
    builder.button(text="Видалити статтю у списку витрат")
    builder.button(text="Відредагувати статтю у списку витрат")
    builder.adjust(2)
    return builder.as_markup(resize_keyboard=True)


@dp.message(CommandStart())
async def start(message: types.Message):
    await message.answer(
        "Ласкаво просимо! Оберіть дію:", reply_markup=create_menu_keyboard()
    )


@dp.message(lambda message: message.text == "Додати статтю витрат")
async def add_expense_start(message: types.Message, state: FSMContext):
    await message.answer("Введіть назву статті витрат:")
    await state.set_state(ExpenseState.name)


@dp.message(StateFilter(ExpenseState.name))
async def add_expense_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer("Введіть дату у форматі dd.mm.YYYY:")
    await state.set_state(ExpenseState.date)


@dp.message(StateFilter(ExpenseState.date))
async def add_expense_date(message: types.Message, state: FSMContext):
    try:
        datetime.datetime.strptime(message.text, "%d.%m.%Y")
        await state.update_data(date=message.text)
        await message.answer("Введіть суму витрат:")
        await state.set_state(ExpenseState.amount)
    except ValueError:
        await message.answer("Невірний формат дати. Спробуйте ще раз.")


@dp.message(StateFilter(ExpenseState.amount))
async def add_expense_amount(message: types.Message, state: FSMContext):
    try:
        amount = float(message.text)
        data = await state.get_data()
        expense = {
            "name": data["name"],
            "date": data["date"],
            "amount": amount,
        }
        async with aiohttp.ClientSession() as session:
            result = await add_expense(session, expense)
            await message.answer(result["message"], reply_markup=create_menu_keyboard())
        await state.clear()
    except ValueError:
        await message.answer("Невірний формат суми. Спробуйте ще раз.")


@dp.message(lambda message: message.text == "Отримати звіт витрат за період")
async def get_report_start(message: types.Message, state: FSMContext):
    await message.answer("Введіть дату початку періоду у форматі dd.mm.YYYY:")
    await state.set_state(ReportState.start_date)


@dp.message(StateFilter(ReportState.start_date))
async def get_report_start_date(message: types.Message, state: FSMContext):
    try:
        datetime.datetime.strptime(message.text, "%d.%m.%Y")
        await state.update_data(start_date=message.text)
        await message.answer("Введіть дату кінця періоду у форматі dd.mm.YYYY:")
        await state.set_state(ReportState.end_date)
    except ValueError:
        await message.answer("Невірний формат дати. Спробуйте ще раз.")


@dp.message(StateFilter(ReportState.end_date))
async def get_report_end_date(message: types.Message, state: FSMContext):
    try:
        datetime.datetime.strptime(message.text, "%d.%m.%Y")
        data = await state.get_data()
        async with aiohttp.ClientSession() as session:
            result = await get_report(
                session, data["start_date"], message.text
            )
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["ID", "Назва", "Дата", "Сума"])
            total_amount = 0
            for expense in result["expenses"]:
                sheet.append(
                    [
                        expense["id"],
                        expense["name"],
                        expense["date"],
                        expense["amount"],
                    ]
                )
                total_amount += expense["amount"]
            file_name = f"{uuid.uuid4()}.xlsx"
            wb.save(file_name)
            with open(file_name, "rb") as file:
                await message.answer_document(
                    types.BufferedInputFile(file.read(), filename=file_name),
                    caption=f"Загальна сума витрат: {total_amount}",
                    reply_markup=create_menu_keyboard(),
                )
            os.remove(file_name)
        await state.clear()
    except ValueError:
        await message.answer("Невірний формат дати. Спробуйте ще раз.")


@dp.message(lambda message: message.text == "Видалити статтю у списку витрат")
async def delete_expense_start(message: types.Message, state: FSMContext):
    async with aiohttp.ClientSession() as session:
        expenses = await get_expenses(session)
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["ID", "Назва", "Дата", "Сума"])
        for expense in expenses:
            sheet.append(
                [expense["id"], expense["name"], expense["date"], expense["amount"]]
            )
        file_name = f"{uuid.uuid4()}.xlsx"
        wb.save(file_name)
        with open(file_name, "rb") as file:
            await message.answer_document(
                types.BufferedInputFile(file.read(), filename=file_name),
                caption="Оберіть ID статті витрат для видалення:",
            )
        os.remove(file_name)
    await state.set_state(DeleteExpenseState.expense_id)


@dp.message(StateFilter(DeleteExpenseState.expense_id))
async def delete_expense_id(message: types.Message, state: FSMContext):
    try:
        expense_id = int(message.text)
        async with aiohttp.ClientSession() as session:
            result = await delete_expense(session, expense_id)
            await message.answer(result["message"], reply_markup=create_menu_keyboard())
        await state.clear()
    except ValueError:
        await message.answer("Невірний формат ID. Спробуйте ще раз.")


@dp.message(lambda message: message.text == "Відредагувати статтю у списку витрат")
async def edit_expense_start(message: types.Message, state: FSMContext):
    async with aiohttp.ClientSession() as session:
        expenses = await get_expenses(session)
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["ID", "Назва", "Дата", "Сума"])
        for expense in expenses:
            sheet.append(
                [expense["id"], expense["name"], expense["date"], expense["amount"]]
            )
        file_name = f"{uuid.uuid4()}.xlsx"
        wb.save(file_name)
        with open(file_name, "rb") as file:
            await message.answer_document(
                types.BufferedInputFile(file.read(), filename=file_name),
                caption="Оберіть ID статті витрат для редагування:",
            )
        os.remove(file_name)
        await state.set_state(EditExpenseState.expense_id)


@dp.message(StateFilter(EditExpenseState.expense_id))
async def edit_expense_id(message: types.Message, state: FSMContext):
    try:
        expense_id = int(message.text)
        await state.update_data(expense_id=expense_id)
        async with aiohttp.ClientSession() as session:
            expenses = await get_expenses(session)
            expense_to_edit = next(
                (expense for expense in expenses if expense["id"] == expense_id), None
            )
            if expense_to_edit:
                await message.answer(
                    f"Назва: {expense_to_edit['name']}\nСума: {expense_to_edit['amount']}\nВведіть нову назву:"
                )
                await state.set_state(EditExpenseState.new_name)
            else:
                await message.answer(
                    "Стаття витрат з таким ID не знайдена.",
                    reply_markup=create_menu_keyboard(),
                )
    except ValueError:
        await message.answer("Невірний формат ID. Спробуйте ще раз.")


@dp.message(StateFilter(EditExpenseState.new_name))
async def edit_expense_new_name(message: types.Message, state: FSMContext):
    await state.update_data(new_name=message.text)
    await message.answer("Введіть нову суму витрат:")
    await state.set_state(EditExpenseState.new_amount)


@dp.message(StateFilter(EditExpenseState.new_amount))
async def edit_expense_new_amount(message: types.Message, state: FSMContext):
    try:
        new_amount = float(message.text)
        data = await state.get_data()
        async with aiohttp.ClientSession() as session:
            result = await update_expense(
                session,
                data["expense_id"],
                {"name": data["new_name"], "amount": new_amount},
            )
            await message.answer(result["message"], reply_markup=create_menu_keyboard())
        await state.clear()
    except ValueError:
        await message.answer("Невірний формат суми. Спробуйте ще раз.")


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())